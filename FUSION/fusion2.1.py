import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import schedule
from time import sleep
from tqdm import tqdm
import argparse

timestamp_dir = 'C:\\Users\\asus\\Desktop\\WORK\\Timestamps'
os.makedirs(timestamp_dir, exist_ok=True)

def process_csv_files(directory, timestamp_file):
    csv_files = [file for file in os.listdir(directory) if file.endswith('.csv')]

    k_field_mapping = {
    'K1001': 'Part_Number',
    'K1002': 'PartTitle',
    'K1003': 'Aggregate',
    'K1004': 'Part Amendment status',
    'K1005': 'Component',
    'K1007': 'Part number â€“ Abbreviation',
    'K1008': 'Model',
    'K1009': 'Part code',
    'K1011': 'Variant',
    'K1022': 'Manufacturer name',
    'K1041': 'Drawing number',
    'K1042': 'Drawing Amendment',
    'K1053': 'Contract',
    'K1072': 'Supplier Description',
    'K1081': 'Machine Number',
    'K1082': 'Machine Description',
    'K1083': 'Machine Number',
    'K1085': 'Machine Location',
    'K1086': 'Operation',
    'K1087': 'Work Cycle Description',
    'K1100': 'Plant',
    'K1101': 'Department',
    'K1102': 'Gaging-Station_Name',
    'K1103': 'Cost centre',
    'K1110': 'Order number',
    'K1201': 'Test Facility Number',
    'K1202': 'Test Facility Description',
    'K1203': 'Reason for Test',
    'K1206': 'Test Location',
    'K1209': 'Inspection type',
    'K1230': 'Gage room',
    'K2001': 'Characteristic Number',
    'K2002': 'Characteristic Description',
    'K2003': 'Characteristic Abbreviation',
    'K2004': 'Characteristic Type',
    'K2005': 'Characteristics Class',
    'K2006': 'Control Item',
    'K2007': 'Control Type',
    'K2008': 'Group type',
    'K2009': 'Measured quantity',
    'K2015': 'Tool wear type (Trend)',
    'K2016': '100% Measurement',
    'K2022': 'Decimal Places',
    'K2043': 'Measuring Device Name',
    'K2060': 'Events Catalo',
    'K2061': 'Process Parameter',
    'K2062': 'Cavity catalogue',
    'K2063': 'Machine catalogue',
    'K2064': 'Gage catalogue',
    'K2065': 'Operator catalogue',
    'K2066': 'Sub-catalogue K0061',
    'K2067': 'Sub-catalogue K0062',
    'K2068': 'Sub-catalogue K0063',
    'K2092': 'Characteristic text',
    'K2093': 'Processing status',
    'K2100': 'Target value',
    'K2101': 'Nominal Value',
    'K2110': 'Lower Specification Limit',
    'K2111': 'Upper Specification Limit',
    'K2112': 'Lower Allowance',
    'K2113': 'Upper Allowance',
    'K2114': 'Lower Scrap Limit',
    'K2115': 'Upper Scrap Limit',
    'K2120': 'Lower Boundary Type',
    'K2121': 'Upper Boundary Type',
    'K2130': 'Lower Plausibility Limit',
    'K2131': 'Upper Plausibility Limit',
    'K2142': 'Unit',
    'K2281': 'Calibration Part Number middle',
    'K2301': 'Machine number',
    'K2302': 'Machine Description',
    'K2303': 'Department/Cost centre',
    'K2311': 'Production Type (Operation)',
    'K2312': 'Production Type Description',
    'K2320': 'Contract Number',
    'K2401': 'Gage Number',
    'K2402': 'Gage Description',
    'K2403': 'Gage Group',
    'K2404': 'Gage Resolution',
    'K0001': 'Measured value',
    'K0002': 'Attribute',
    'K0004': 'Date/Time',
    'K0005': 'Event',
    'K0006': 'Batch number',
    'K0007': 'Cavity number',
    'K0008': 'Operator name',
    'K0009': 'Text',
    'K0010': 'Machine number',
    'K0011': 'Process parameter',
    'K0012': 'Gage number',
    'K0014': 'Part ID',
    'K0015': 'Reason for test',
    'K0016': 'Production number',
    'K0017': 'Work piece fixture number',
    'K0053': 'Order',
    'K0054': 'K0054',
    'K0055': 'K0055',
    'K0056': 'K0056',
    'K0057': 'K0057',
    'K0058': 'K0058',
    'K0059': 'K0059',
    'K0060': 'K0060',
    'K0061': 'K0061',
    'K0062': 'K0062',
    'K0063': 'K0063'
    }

    
    last_processed_timestamps = {}
    if os.path.exists(timestamp_file):
        with open(timestamp_file, 'r') as f:
            lines = f.readlines()
            for line in lines:
                filename, timestamp = line.strip().split(',')
                last_processed_timestamps[filename] = float(timestamp)
    
    for csv_file in csv_files:
        csv_file_path = os.path.join(directory, csv_file)
        file_timestamp = os.path.getmtime(csv_file_path)
        
        if csv_file not in last_processed_timestamps or file_timestamp > last_processed_timestamps[csv_file]:
            with open(csv_file_path, 'r', encoding='cp1252') as file:
                lines = file.readlines()
            
            split_index = -1
            for i, line in enumerate(lines):
                if line.startswith("Measurement Name"):
                    split_index = i-1
                    break
            
            if split_index != -1:
                part_1_lines = lines[:split_index + 1]
                csv_1 = "csv_1.csv"
                with open(csv_1, 'w', encoding='cp1252') as file:
                    file.writelines(part_1_lines)

                part_2_lines = lines[split_index + 1:]
                csv_2 = "csv_2.csv"
                with open(csv_2, 'w', encoding='cp1252') as file:
                    file.writelines(part_2_lines)

            df_part_1 = pd.read_csv(csv_1, sep=',', header=None, encoding='cp1252')
            columns_to_add_1 = 12 - len(df_part_1.iloc[0])
            padding_1 = pd.DataFrame([[0] * columns_to_add_1] * len(df_part_1))
            df_padded_1 = pd.concat([df_part_1, padding_1], axis=1)

            excel_file_1 = "output_part_1.xlsx"
            workbook_1 = Workbook()
            sheet_1 = workbook_1.active

            for r_idx, row in enumerate(df_padded_1.values):
                for c_idx, value in enumerate(row):
                    sheet_1.cell(row=r_idx + 1, column=c_idx + 1, value=value)

            workbook_1.save(excel_file_1)

            df_part_2 = pd.read_csv(csv_2, sep=',', header=None, encoding='cp1252')
            columns_to_add_2 = 12 - len(df_part_2.iloc[0])
            padding_2 = pd.DataFrame([[0] * columns_to_add_2] * len(df_part_2))
            df_padded_2 = pd.concat([df_part_2, padding_2], axis=1)

            excel_file_2 = "output_part_2.xlsx"
            workbook_2 = Workbook()
            sheet_2 = workbook_2.active

            for r_idx, row in enumerate(df_padded_2.values):
                for c_idx, value in enumerate(row):
                    sheet_2.cell(row=r_idx + 1, column=c_idx + 1, value=value)

            workbook_2.save(excel_file_2)

            loaded_workbook_2 = load_workbook(excel_file_2)
            loaded_sheet_2 = loaded_workbook_2.active
            row_count = loaded_sheet_2.max_row

            loaded_workbook_1 = load_workbook(excel_file_1)
            loaded_sheet_1 = loaded_workbook_1.active

            def sciNotation(value):
                scientific_notation = "{:.14E}".format(value).replace('e', 'E').replace('+', '+00').replace('-', '-00')
                coefficient, exponent = scientific_notation.split('E')
                formatted_exponent = exponent[:1] + exponent[1:].rjust(4, '0')
                formatted_scientific_notation = f"{coefficient}E{formatted_exponent}"
                return formatted_scientific_notation

            dateTime = loaded_sheet_1["C8"].value
            parsed_datetime = datetime.strptime(dateTime, "%d-%b-%Y %H:%M:%S")
            formatted_date = parsed_datetime.strftime("%d.%m.%Y/%H:%M:%S")

        # Generate the output DFQ file name based on the original CSV file name
            dfq_file_name = os.path.splitext(csv_file)[0] + ".dfq"
        
        # Specify the path for the output DFQ file in the output folder
            output_dfq_file_path = os.path.join(custom_output_folder, dfq_file_name)

            with open(output_dfq_file_path, "w") as output_file:
                output_file.write(f'K0100 {(row_count - 1)}\n')
                output_file.write('K0101 2\n')

                for k_field, column_name in k_field_mapping.items():
                    row_number = None

                    for row_number, row in enumerate(loaded_sheet_1.iter_rows(), start=1):
                        for cell in row:
                            if cell.value == column_name:
                                output_file.write(f'{k_field}/1 {loaded_sheet_1[f"C{row_number}"].value}\n')
                                break
                        else:
                            continue
                        break 

            # with open(output_dfq_file_path, "w") as output_file:
            #     output_file.write(f'K0100 {(row_count - 1)}\n')
            #     output_file.write('K0101 2\n')
            
            #     output_file.write(f'K1001/1 {loaded_sheet_1["C20"].value}\n')
            #     output_file.write(f'K1002/1 {loaded_sheet_1["C1"].value}\n')
            #     output_file.write(f'K1003/1 {loaded_sheet_1["C21"].value}\n')
            #     output_file.write(f'K1005/1 {loaded_sheet_1["C23"].value}\n')
            #     output_file.write(f'K1008/1 {loaded_sheet_1["C22"].value}\n')
            #     output_file.write(f'K1086/1 {loaded_sheet_1["C24"].value}\n')
            #     output_file.write(f'K1100/1 {loaded_sheet_1["C25"].value}\n')
            #     output_file.write(f'K1102/1 {loaded_sheet_1["C19"].value}\n')

                for i in range(1, row_count):
                    output_file.write(f'K2001/{i} {i}\n')
                    output_file.write(f'K2002/{i} {loaded_sheet_2[f"A{i+1}"].value}\n')
                    output_file.write(f'K2101/{i} {loaded_sheet_2[f"E{i+1}"].value}\n')
                    output_file.write(f'K2110/{i} {loaded_sheet_2[f"G{i+1}"].value}\n')
                    output_file.write(f'K2111/{i} {loaded_sheet_2[f"F{i+1}"].value}\n')

                pallet_id = loaded_sheet_1["C27"].value

                for i in range(1, row_count):
                    if(i!=row_count):
                        cell_value = loaded_sheet_2[f"C{i+1}"].value
                        if cell_value is not None and cell_value.strip() != "":
                            try:
                                value = float(cell_value)
                            except ValueError:
                                value = 0.0  # Handle non-numeric values gracefully
                        else:
                            value = 0.0 
                    output_file.write(f'{sciNotation(value)}0{formatted_date}#{pallet_id}0000')
                    output_file.write('\n')
        
                for i in range(1, row_count):
                    output_file.write(f'K0053/{i} {loaded_sheet_1["C28"].value}\n')
                    output_file.write(f'K0014/{i} {loaded_sheet_1["C10"].value}\n')
                    output_file.write(f'K0054/{i} {loaded_sheet_1["C26"].value}\n')

            os.remove(csv_1)
            os.remove(csv_2)
            os.remove(excel_file_1)
            os.remove(excel_file_2)
            
            # Update the timestamp for the processed file in the dictionary
            last_processed_timestamps[csv_file] = file_timestamp
    
    # Save the updated timestamps to the timestamp file
    with open(timestamp_file, 'w') as f:
        for filename, timestamp in last_processed_timestamps.items():
            f.write(f'{filename},{timestamp}\n')

def job(root_directory):
    print(f"Running the converter script in {root_directory} and its subdirectories...")
    for i in tqdm(range(0, 10), colour="#B0Fc38", desc="Initializing..."):
        sleep(.1)
    for root, dirs, files in os.walk(root_directory):
        for dir in dirs:
            subdirectory_path = os.path.join(root, dir)
            print(subdirectory_path)
            timestamp_file = os.path.join(timestamp_dir, f'{dir}_timestamps.txt')
            process_csv_files(subdirectory_path, timestamp_file)

    for i in tqdm(range(0, 10), colour="#03C04A", desc="Progress: "):
        sleep(.1)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="CSV -> DFQ Converter")
    directory = input("Specify the root directory to process CSV files: ")
    #parser.add_argument("directory", help="Specify the root directory to process CSV files")
    #args = parser.parse_args()

    custom_output_folder = 'C:\\Users\\asus\\Desktop\\WORK\\OUTPUT_DFQ'
    os.makedirs(custom_output_folder, exist_ok=True)

    print(f"Initializing the converter for directory: {directory}")
    job(directory)

    schedule.every(20).seconds.do(job, directory)
    
    while True:
        schedule.run_pending()
        sleep(1)

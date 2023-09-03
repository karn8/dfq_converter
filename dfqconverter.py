import pandas as pd
import os
from openpyxl import Workbook, load_workbook
from datetime import datetime

# store the path to the csv file in csv_file variable 
csv_file = 'C:\\Users\\asus\\Desktop\\WORK\\084.csv'
# csv_file = input()
with open(csv_file, 'r', encoding='cp1252') as file:
    lines = file.readlines()

part_1_lines = lines[:28]
csv_1 = "csv_1.csv"
with open(csv_1, 'w', encoding='cp1252') as file:
    file.writelines(part_1_lines)

part_2_lines = lines[28:]
csv_2 = "csv_2.csv"
with open(csv_2, 'w', encoding='cp1252') as file:
    file.writelines(part_2_lines)

# csv_1 to dataframe
df_part_1 = pd.read_csv(csv_1, sep=',', header=None, encoding = 'cp1252')
columns_to_add_1 = 12 - len(df_part_1.iloc[0])
padding_1 = pd.DataFrame([[0] * columns_to_add_1] * len(df_part_1))
df_padded_1 = pd.concat([df_part_1, padding_1], axis=1)

excel_file_1 = "output_part_1.xlsx"
workbook_1 = Workbook()
sheet_1 = workbook_1.active

for r_idx, row in enumerate(df_padded_1.values):
    for c_idx, value in enumerate(row):
        sheet_1.cell(row=r_idx + 1, column=c_idx + 1, value=value)

workbook_1.save(excel_file_1)

# csv_2 to dataframe
df_part_2 = pd.read_csv(csv_2, sep=',', header=None, encoding = 'cp1252')
columns_to_add_2 = 12 - len(df_part_2.iloc[0])
padding_2 = pd.DataFrame([[0] * columns_to_add_2] * len(df_part_2))
df_padded_2 = pd.concat([df_part_2, padding_2], axis=1)

excel_file_2 = "output_part_2.xlsx"
workbook_2 = Workbook()
sheet_2 = workbook_2.active

for r_idx, row in enumerate(df_padded_2.values):
    for c_idx, value in enumerate(row):
        sheet_2.cell(row=r_idx + 1, column=c_idx + 1, value=value)

workbook_2.save(excel_file_2)

loaded_workbook_2 = load_workbook(excel_file_2)
loaded_sheet_2 = loaded_workbook_2.active
row_count = loaded_sheet_2.max_row

loaded_workbook_1 = load_workbook(excel_file_1)
loaded_sheet_1 = loaded_workbook_1.active

# Fixed Width scientific notation
def sciNotation(value):
    scientific_notation = "{:.14E}".format(value).replace('e', 'E').replace('+', '+00').replace('-', '-00')
    coefficient, exponent = scientific_notation.split('E')
    formatted_exponent = exponent[:1] + exponent[1:].rjust(4, '0')
    formatted_scientific_notation = f"{coefficient}E{formatted_exponent}"
    return formatted_scientific_notation

dateTime = loaded_sheet_1["C8"].value

# Parse input Date/Time
parsed_datetime = datetime.strptime(dateTime, "%d-%b-%Y %H:%M:%S")
formatted_date = parsed_datetime.strftime("%d.%m.%Y/%H:%M:%S")

with open(f"{loaded_sheet_1['C10'].value}.dfq", "w") as output_file:
    output_file.write(f'K0100 {(row_count - 1)}\n')
    output_file.write('K0101 2\n')
    
    # Part Details
    output_file.write(f'K1001/1 {loaded_sheet_1["C20"].value}\n')
    output_file.write(f'K1002/1 {loaded_sheet_1["C1"].value}\n')
    output_file.write(f'K1003/1 {loaded_sheet_1["C21"].value}\n')
    output_file.write(f'K1005/1 {loaded_sheet_1["C23"].value}\n')
    output_file.write(f'K1008/1 {loaded_sheet_1["C22"].value}\n')
    output_file.write(f'K1086/1 {loaded_sheet_1["C24"].value}\n')
    output_file.write(f'K1100/1 {loaded_sheet_1["C25"].value}\n')
    output_file.write(f'K1102/1 {loaded_sheet_1["C19"].value}\n')

    for i in range(1, 131):
        output_file.write(f'K2001/{i} {i}\n')
        output_file.write(f'K2002/{i} {loaded_sheet_2[f"A{i+1}"].value}\n')
        output_file.write(f'K2101/{i} {loaded_sheet_2[f"E{i+1}"].value}\n')
        output_file.write(f'K2110/{i} {loaded_sheet_2[f"G{i+1}"].value}\n')
        output_file.write(f'K2111/{i} {loaded_sheet_2[f"F{i+1}"].value}\n')

    pallet_id = loaded_sheet_1["C27"].value

    # value = float(loaded_sheet_2[f"C{i+1}"].value)
    # print(f'{sciNotation(value)}0{formatted_date}#{pallet_id}0000\n')

    for i in range(1, 131):
        if(i!=131):
            value = float(loaded_sheet_2[f"C{i+1}"].value)
            output_file.write(f'{sciNotation(value)}0{formatted_date}#{pallet_id}0000')
    output_file.write('\n')

    for i in range(1, 131):
        output_file.write(f'K0053/{i} {loaded_sheet_1["C28"].value}\n')
        output_file.write(f'K0014/{i} {loaded_sheet_1["C10"].value}\n')
        output_file.write(f'K0054/{i} {loaded_sheet_1["C26"].value}\n')

# Delete intermediate files
os.remove(csv_1)
os.remove(csv_2)
os.remove(excel_file_1)
os.remove(excel_file_2)